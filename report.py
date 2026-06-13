"""Write match results to a color-coded Excel report."""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

from models import MatchResult

STATUS_TEXT = {
    "OK": "OK",
    "WARNING": "rozdíl",
    "MISSING": "chybí na faktuře",
    "EXTRA": "navíc na faktuře",
}
_FILLS = {
    "OK": PatternFill("solid", start_color="C6EFCE"),       # green
    "WARNING": PatternFill("solid", start_color="FFEB9C"),  # orange
    "MISSING": PatternFill("solid", start_color="FFC7CE"),  # red
    "EXTRA": PatternFill("solid", start_color="FFC7CE"),    # red
}

HEADERS = ["Stav", "Objekt", "Pozice", "Rozměr objednávka", "Rozměr faktura",
           "Ks objednáno", "Ks fakturováno", "Skladba objednávka",
           "Skladba faktura", "Problémy"]


def result_row(r: MatchResult) -> list[str]:
    o, i = r.order_item, r.invoice_item
    return [
        STATUS_TEXT[r.status],
        o.objekt if o else "",
        o.label if o else (i.label if i else ""),
        f"{o.width} x {o.height}" if o else "",
        f"{i.width} x {i.height}" if i else "",
        str(o.quantity) if o else "",
        str(i.quantity) if i else "",
        o.skladba_raw if o else "",
        f"{i.composition_raw} | rámeček {i.spacer} mm" if i else "",
        "; ".join(r.problems),
    ]


def write_report(results: list[MatchResult], path: str | Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kontrola"
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for r in results:
        ws.append(result_row(r))
        fill = _FILLS[r.status]
        for cell in ws[ws.max_row]:
            cell.fill = fill
    for col, width in zip("ABCDEFGHIJ", (16, 16, 10, 16, 16, 12, 12, 24, 40, 50)):
        ws.column_dimensions[col].width = width
    wb.save(path)
